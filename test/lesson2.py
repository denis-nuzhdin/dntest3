from openpyxl import load_workbook

workbook = load_workbook('users.xlsx')
first_sheet = workbook.get_sheet_names()[0]
worksheet = workbook.get_sheet_by_name(first_sheet)

for row in worksheet.iter_rows():
    print (row)

# check out the last row
for cell in row:
    print (cell)